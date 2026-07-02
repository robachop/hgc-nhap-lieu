"""
Sinh kế hoạch đảo trộn Mien từ sheet "Tuần XX" trong file Excel
"0. HGC Sổ làm việc - Họp đầu và cuối ca.xlsx" (hoặc file cùng cấu trúc).

Cấu trúc sheet: cột A=STT, B=BỂ (số bể thô), C=CK (chu kỳ 1-6), bắt đầu từ dòng 4.
Cùng 1 danh sách bể lặp lại mỗi ngày làm việc trong tuần (đã Tim xác nhận 2026-07-02).

Form chuẩn LSX (Tim chốt 2026-07-02):
  lsx     = S{ck}{day_idx:02d}   (day_idx = thứ tự ngày trong tuần, 0-based:
                                   ngày đầu tuần -> 00, ngày kế -> 01, ngày kế nữa -> 02...
                                   vd CK6: 06/07->S600, 07/07->S601, 08/07->S602...)
  be_cap  = T{bể, pad 3 số}      (bể cấp/nguồn, vd bể 148 -> T148, bể 8 -> T008)
  be_nhan = L{bể, pad 3 số}      (bể nhận, vd bể 148 -> L148, bể 8 -> L008)

Dùng:
  python3 scripts/gen_mien_tuan.py "<file.xlsx>" "Tuần 28" 2026-07-06 6

  <file.xlsx>   đường dẫn Excel
  "Tuần XX"     tên sheet chứa danh sách bể/CK
  YYYY-MM-DD    ngày đầu tuần (thứ 2)
  N             số ngày làm việc liên tiếp (mặc định 6, bỏ CN tự động nếu rơi vào)
"""
import sys, json, datetime
import openpyxl

# Đối chiếu S500 (2026-07-02) — Tim đã xác nhận 4 nhóm nghi vấn của Tuần 28:
#   1. NHẢY CÓC chu kỳ (Tim: "Đồng ý" = sửa về đúng chu kỳ kế tiếp so với S500)
CK_OVERRIDE = {
    167: 5,   # S500 cao nhất CK4 -> đúng ra CK5, sheet gốc ghi nhầm CK6
    168: 4,   # S500 cao nhất CK3 -> đúng ra CK4, sheet gốc ghi nhầm CK5
    169: 4,   # S500 cao nhất CK3 -> đúng ra CK4, sheet gốc ghi nhầm CK5
}
#   2. Đã hết hạn ngày chu kỳ hiện tại nhưng sheet gốc lặp lại CK cũ
#      (Tim: "ghi là trùng lặp" — giữ nguyên CK, chỉ đánh dấu trong mô tả)
DUPLICATE_NOTE_BE = {163, 188}
#   3. 7 bể quay lại CK1 sau khi đã xong CK7 (Tim: "đã qua chượp mới, không phải lỗi") -> không sửa gì
#   4. Bể 354 không có lịch sử trong S500 (Tim: "lỗi đánh máy sai", xác nhận số đúng là 35)
BE_NUMBER_FIX = {354: 35}

def read_be_ck(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    r = 4
    while True:
        be = ws.cell(r, 2).value
        ck = ws.cell(r, 3).value
        stt = ws.cell(r, 1).value
        if stt is None and be is None:
            break
        if be is not None and ck is not None:
            if be in BE_NUMBER_FIX:
                be = BE_NUMBER_FIX[be]
            if be in CK_OVERRIDE:
                ck = CK_OVERRIDE[be]
            rows.append((be, ck))
        r += 1
    return rows

def build_tasks(rows, day_idx):
    tasks = []
    for i, (be, ck) in enumerate(rows, start=1):
        mo_ta = f"Đảo trộn bể {be} (CK{ck})"
        if be in DUPLICATE_NOTE_BE:
            mo_ta += " ⚠️ TRÙNG LẶP (đã hết hạn CK cũ trong S500)"
        tasks.append({
            "id": f"t{i}",
            "nguoi": "Mien",
            "lsx": f"S{ck}{day_idx:02d}",
            "mo_ta": mo_ta,
            "be_cap": f"T{be:03d}",
            "be_nhan": f"L{be:03d}",
            "luong_dk": 0,
            "dvt": "lít",
            "cong": 5,
            "group": "S_dao_tron"
        })
    return tasks

def next_workday(d):
    while d.weekday() == 6:  # bỏ Chủ Nhật
        d += datetime.timedelta(days=1)
    return d

def main():
    excel_path, sheet_name, start_str, n_days = sys.argv[1], sys.argv[2], sys.argv[3], int(sys.argv[4]) if len(sys.argv) > 4 else 6
    rows = read_be_ck(excel_path, sheet_name)
    print(f"Đọc {len(rows)} bể từ sheet '{sheet_name}'")

    d = datetime.date.fromisoformat(start_str)
    days = []
    while len(days) < n_days:
        d = next_workday(d)
        days.append(d)
        d += datetime.timedelta(days=1)

    for day_idx, d in enumerate(days):
        slug = d.strftime('%d%m%Y')
        plan = {"date": d.isoformat(), "tasks": build_tasks(rows, day_idx)}
        json.dump(plan, open(f'plans/mien-{slug}.json', 'w'), ensure_ascii=False, indent=2)

        redirect_url = f"https://robachop.github.io/hgc-nhap-lieu/?plan_file=mien-{slug}&w=Mien"
        html = f'''<!DOCTYPE html>
<html lang="vi"><head><meta charset="UTF-8">
<meta http-equiv="refresh" content="0;url={redirect_url}">
<title>HGC Kế Hoạch Mien — {d.strftime('%d/%m/%Y')}</title>
</head><body>
<p>Đang chuyển hướng... <a href="{redirect_url}">Bấm đây nếu không tự chuyển</a></p>
<script>window.location.replace("{redirect_url}");</script>
</body></html>'''
        open(f'kehoach-mien-{slug}.html', 'w').write(html)

    print("Đã tạo:", [d.strftime('%d/%m') for d in days])

if __name__ == '__main__':
    main()
