#!/usr/bin/env python3
"""
tao_ke_hoach.py — Tạo kế hoạch tự động từ S500 và deploy lên PWA.

Usage:
    python3 tao_ke_hoach.py <file_excel.xlsx> [YYYY-MM-DD]
    Nếu không truyền ngày → mặc định ngày làm việc tiếp theo (bỏ qua CN).
"""

import sys, json, re, base64, datetime, subprocess
from pathlib import Path
from collections import defaultdict, Counter
from urllib.parse import quote

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# ── Config ────────────────────────────────────────────────────
WORKERS = ['Ha', 'Hao', 'Mien', 'Phong', 'Toan']
REPO_DIR = Path(__file__).parent.parent   # /tmp/hgc-nhap-lieu
BASE_URL = "https://robachop.github.io/hgc-nhap-lieu/"

# Vai chuyên môn (để filter đúng nhóm LSX cho từng người)
ROLES = {
    'Ha':    ['C040','C_keo_rut','N_cot_nhi'],
    'Hao':   ['C040','N_cot_nhi','PM00'],
    'Mien':  ['S_series','S_dao_tron'],
    'Phong': ['C040','C_rut_cot','PM00','P_xuat','PX_rut_kiet'],
    'Toan':  ['C040','S_dao_tron'],
}

# ── Helpers ───────────────────────────────────────────────────
def days_ago(d, ref):
    return (ref - d).days

def next_workday(ref=None):
    if ref is None: ref = datetime.date.today()
    d = ref + datetime.timedelta(days=1)
    while d.weekday() == 6:   # bỏ CN
        d += datetime.timedelta(days=1)
    return d

def encode_plan(plan):
    compact = {
        'd': plan['date'],
        't': [{'i':t['id'],'c':t.get('be_cap',''),'l':t['lsx'],
               'n':t['be_nhan'],'q':t.get('luong_dk',0),
               'p':t['nguoi'],'g':t.get('ghi_chu','')}
              for t in plan['tasks']]
    }
    raw = json.dumps(compact, ensure_ascii=False, separators=(',',':')).encode('utf-8')
    return quote(base64.b64encode(raw).decode(), safe='')

# ── Load data ─────────────────────────────────────────────────
def load_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # LSX info
    lsx_info = {}
    for i, row in enumerate(wb['LSX Mẫu'].iter_rows(values_only=True)):
        if i == 0: continue
        if row[2]:
            lsx_info[str(row[2]).strip()] = {
                'mo_ta': str(row[5] or row[3] or '').strip(),
                'dvt':   str(row[8] or 'lít').strip(),
                'cong':  int(row[9] or 5)
            }

    # S500 — latest per (worker, bể)
    w_latest  = {w: {} for w in WORKERS}
    w_history = {w: defaultdict(list) for w in WORKERS}

    for i, row in enumerate(wb['S500'].iter_rows(values_only=True)):
        if i == 0 or row[0] is None: continue
        nguoi = str(row[9] or '').strip()
        if nguoi not in WORKERS: continue
        ngay = row[7]
        if hasattr(ngay, 'date'): ngay = ngay.date()
        elif not isinstance(ngay, datetime.date): continue
        be    = str(row[5] or '').strip()
        lsx   = str(row[6] or '').strip()
        luong = float(row[8] or 0)
        w_history[nguoi][be].append((ngay, lsx, luong))
        if be not in w_latest[nguoi] or ngay > w_latest[nguoi][be][1]:
            w_latest[nguoi][be] = (lsx, ngay, luong)

    return lsx_info, w_latest, w_history

# ── Plan generation ───────────────────────────────────────────
def make_tasks_for(worker, latest, lsx_info, ref_date):
    """Tạo danh sách task gợi ý cho 1 công nhân."""
    tasks = []
    tid = [0]

    def add(group, lsx, be_nhan, be_cap, luong, ghi_chu, priority='B'):
        tid[0] += 1
        info = lsx_info.get(lsx, {'mo_ta': lsx, 'dvt':'lít', 'cong':5})
        tasks.append({
            'id': f't{tid[0]}',
            'group': group,
            'priority': priority,
            'be_cap':   be_cap,
            'lsx':      lsx,
            'mo_ta':    info['mo_ta'],
            'dvt':      info['dvt'],
            'cong':     info['cong'],
            'be_nhan':  be_nhan,
            'luong_dk': int(luong),
            'nguoi':    worker,
            'ghi_chu':  ghi_chu,
        })

    roles = ROLES.get(worker, [])

    # --- C040 Đảo trong (chu kỳ 3-4 ngày) ---
    if 'C040' in roles:
        c040 = [(be, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                if lsx == 'C040' and be.startswith('L') and 3 <= days_ago(d, ref_date) <= 14]
        for be, d, lu in sorted(c040, key=lambda x: -days_ago(x[1], ref_date))[:8]:
            add('Đảo trong', 'C040', be, '', lu or 2500,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    # --- C-series kéo rút (C1x0..C9x0, done trong 7 ngày) ---
    if 'C_keo_rut' in roles:
        c_keo = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                 if re.match(r'^C[1-9]\d0$', lsx) and days_ago(d, ref_date) <= 7]
        for be, lsx, d, lu in sorted(c_keo, key=lambda x: x[0])[:8]:
            add('Kéo rút', lsx, be, '', lu or 2000,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    # --- C030/C010 (rút cốt gần đây) ---
    if 'C_rut_cot' in roles:
        c_rut = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                 if lsx in ('C030','C010','C020') and days_ago(d, ref_date) <= 5]
        for be, lsx, d, lu in sorted(c_rut, key=lambda x: x[2]):
            add('Rút cốt / Đảo trong', lsx, be, '', lu or 5000,
                f'Lần cuối {d.strftime("%d/%m")} — theo dõi tiếp', 'A')

    # --- S-series: S010/S020 bể trống ---
    if 'S_series' in roles:
        s0 = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
              if re.match(r'^S0[12]0$', lsx) and days_ago(d, ref_date) <= 4]
        for be, lsx, d, lu in sorted(s0, key=lambda x: x[2])[:4]:
            add('Bể trống / Chuẩn bị', lsx, be, '', lu or 40000,
                f'Lần cuối {d.strftime("%d/%m")} — chuẩn bị nhận cá', 'B')

        # S030 Nhập cá
        s030 = [(be, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                if lsx == 'S030' and days_ago(d, ref_date) <= 5]
        for be, d, lu in sorted(s030, key=lambda x: x[1])[:5]:
            add('Nhập cá', 'S030', be, '', lu or 20000,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    # --- S-series đảo trộn ---
    if 'S_dao_tron' in roles:
        s_dt = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                if re.match(r'^S[1-7]\d{2}$', lsx) and 2 <= days_ago(d, ref_date) <= 8]
        for be, lsx, d, lu in sorted(s_dt, key=lambda x: x[2])[:10]:
            add('Đảo trộn', lsx, be, '', lu or 4000,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    # --- PM00 Nước bổi ---
    if 'PM00' in roles:
        pm = [(be, d, int(lu)) for be, (lsx, d, lu) in latest.items()
              if lsx == 'PM00' and 2 <= days_ago(d, ref_date) <= 10]
        for be, d, lu in sorted(pm, key=lambda x: -days_ago(x[1], ref_date))[:5]:
            add('Nước bổi', 'PM00', be, 'Txxx', lu or 3000,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n) — xác nhận Txxx', 'B')

    # --- Cốt nhỉ N-series ---
    if 'N_cot_nhi' in roles:
        n_tasks = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
                   if re.match(r'^N\d+', lsx) and 2 <= days_ago(d, ref_date) <= 10]
        for be, lsx, d, lu in sorted(n_tasks, key=lambda x: -days_ago(x[2], ref_date))[:4]:
            src = 'L'+lsx[1:].zfill(3) if lsx[1:].isdigit() else ''
            add('Cốt nhỉ', lsx, be, src, lu or 1000,
                f'Cốt nhỉ từ {src} → {be} | Lần cuối {d.strftime("%d/%m")}', 'B')

    # --- P-series kéo rút nước long (vào bể TP) ---
    if 'P_xuat' in roles:
        p_tp = defaultdict(list)
        for be, (lsx, d, lu) in latest.items():
            if re.match(r'^P[1-6]\d0$', lsx) and days_ago(d, ref_date) <= 10:
                p_tp[be].append((lsx, d, int(lu)))
        for tp_be, items in sorted(p_tp.items()):
            lsx, d, lu = max(items, key=lambda x: x[1])
            add('Kéo rút nước long', lsx, tp_be, '', lu or 1500,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    # --- PX Rút kiệt ---
    if 'PX_rut_kiet' in roles:
        px = [(be, lsx, d, int(lu)) for be, (lsx, d, lu) in latest.items()
              if re.match(r'^P[TX]\d*', lsx) and days_ago(d, ref_date) <= 8]
        for be, lsx, d, lu in sorted(px, key=lambda x: -days_ago(x[2], ref_date))[:5]:
            add('Xuất / Rút kiệt', lsx, be, '', lu or 13000,
                f'Lần cuối {d.strftime("%d/%m")} ({days_ago(d,ref_date)}n)', 'A')

    tasks.sort(key=lambda x: (x['priority'], x['group']))
    for i, t in enumerate(tasks, 1): t['id'] = f't{i}'
    return tasks

# ── Deploy redirect page ──────────────────────────────────────
def deploy_redirect(plan, encoded_url, target_date):
    date_str  = target_date.strftime("%d/%m/%Y")
    day_vn    = ['Thứ 2','Thứ 3','Thứ 4','Thứ 5','Thứ 6','Thứ 7','Chủ Nhật'][target_date.weekday()]
    slug      = target_date.strftime("%d%m%Y")
    fname     = f"kehoach-{slug}.html"
    fpath     = REPO_DIR / fname

    by_w = Counter(t['nguoi'] for t in plan['tasks'])
    worker_rows = ''.join(
        f'<tr><td>👷 {w}</td><td style="text-align:center;font-weight:700">{by_w.get(w,0)}</td></tr>'
        for w in WORKERS
    )

    html = f"""<!DOCTYPE html>
<html lang="vi"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>HGC Kế hoạch {date_str}</title>
<script>window.location.replace("{encoded_url}");</script>
</head>
<body style="font-family:Arial;text-align:center;padding:32px;color:#1e293b;background:#f1f5f9">
  <div style="background:#fff;border-radius:16px;padding:24px;max-width:360px;margin:0 auto;box-shadow:0 2px 8px rgba(0,0,0,.1)">
    <div style="font-size:40px">📋</div>
    <h2 style="margin:8px 0 4px">HGC Kế hoạch</h2>
    <div style="color:#3b82f6;font-weight:700;font-size:18px">{day_vn}, {date_str}</div>
    <table style="width:100%;margin:16px 0;border-collapse:collapse;font-size:14px">
      <tr style="background:#f8fafc"><th style="padding:8px;text-align:left">Công nhân</th><th>Số task</th></tr>
      {worker_rows}
    </table>
    <a href="{encoded_url}"
      style="display:block;padding:14px;background:#1e40af;color:#fff;border-radius:10px;
             text-decoration:none;font-weight:700;font-size:15px">
      📲 Mở app + Tải kế hoạch
    </a>
    <div style="font-size:11px;color:#94a3b8;margin-top:12px">
      {len(plan['tasks'])} task · Tạo tự động bởi Cod
    </div>
  </div>
</body></html>"""

    fpath.write_text(html, encoding='utf-8')

    # Git commit + push
    result = subprocess.run(
        ['git', 'add', fname, '&&',
         'git', 'commit', '-m', f'Kế hoạch {date_str}: {len(plan["tasks"])} task tự động',
         '&&', 'git', 'push', 'origin', 'main'],
        cwd=REPO_DIR, shell=False, capture_output=True, text=True
    )
    # Try with shell=True for && chaining
    n_tasks = len(plan['tasks'])
    msg = f'Ke hoach {date_str}: {n_tasks} task'
    subprocess.run(
        f'cd "{REPO_DIR}" && git add {fname} && git commit -m "{msg}" && git push origin main',
        shell=True, capture_output=True
    )
    return BASE_URL + fname

# ── Main ──────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        # Auto-find latest Excel in Downloads
        downloads = Path.home() / 'Downloads'
        candidates = sorted(downloads.glob('*Phân công*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            print("Không tìm thấy file Excel. Truyền đường dẫn: python3 tao_ke_hoach.py <file.xlsx>")
            sys.exit(1)
        excel_path = candidates[0]
        print(f"📂 Dùng file mới nhất: {excel_path.name}")
    else:
        excel_path = Path(sys.argv[1])

    # Target date
    if len(sys.argv) >= 3:
        target = datetime.date.fromisoformat(sys.argv[2])
    else:
        target = next_workday()

    today = datetime.date.today()
    print(f"📅 Lên kế hoạch cho: {target.strftime('%A %d/%m/%Y')}")
    print(f"📊 Đọc S500...")

    lsx_info, w_latest, w_history = load_excel(excel_path)

    # Generate tasks for all workers
    all_tasks = []
    by_worker = {}
    for w in WORKERS:
        tasks = make_tasks_for(w, w_latest[w], lsx_info, today)
        # Fix IDs to be unique across workers
        offset = len(all_tasks)
        for i, t in enumerate(tasks, offset + 1):
            t['id'] = f't{i}'
        by_worker[w] = tasks
        all_tasks.extend(tasks)

    plan = {'date': target.isoformat(), 'tasks': all_tasks}

    # Encode for PWA
    encoded_plan = encode_plan(plan)
    pwa_url = BASE_URL + f"?plan={encoded_plan}"

    # ── Summary for Tim ────────────────────────────────────────
    day_vn = ['Thứ 2','Thứ 3','Thứ 4','Thứ 5','Thứ 6','Thứ 7','CN'][target.weekday()]
    print(f"\n{'═'*55}")
    print(f"  KẾ HOẠCH GỢI Ý — {day_vn.upper()} {target.strftime('%d/%m/%Y')}")
    print(f"{'═'*55}")

    for w in WORKERS:
        tasks = by_worker[w]
        if not tasks:
            print(f"  {w:8}: ⚠️  Không có task (cần Tim gán tay)")
            continue
        groups = Counter(t['group'] for t in tasks)
        print(f"  {w:8}: {len(tasks):2} task — " + ', '.join(f"{g}×{n}" for g,n in groups.most_common()))

    print(f"\n  Tổng: {len(all_tasks)} task cho {sum(1 for w in WORKERS if by_worker[w])} người")

    # Deploy redirect page
    print(f"\n🚀 Deploy lên GitHub Pages...")
    short_url = deploy_redirect(plan, pwa_url, target)
    slug = target.strftime("%d%m%Y")
    print(f"\n✅ XONG! Link ngắn cho Tim:")
    print(f"   {short_url}")
    print(f"\n📌 Tim mở link này → xem kế hoạch → gửi link từng người qua Zalo")

    # Save JSON backup
    out = Path.home() / 'Downloads' / f"HGC_KEHOACH_{slug}.json"
    out.write_text(json.dumps(plan, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    print(f"📁 Backup JSON: {out.name}")

if __name__ == '__main__':
    main()
